# coding: utf8
from django.db import connection
from openpyxl import Workbook
import datetime
import os

cmd_dict = {'receive_sample': """select *
          from send_sample ss inner join sample_project_master m on m.id=ss.project_id""",
            'quality_check': """select d.id, d.sample_id, d.sample_name,
          d.rin, d.concentration, d.volume, d.qualitycheck_results, qualitycheck_time, qualitycheck_comment
          from sample_info_detail d inner join sample_project_master m on m.id=d.project_id""",
            'build_lib': """select d.id, d.sample_id, d.sample_name,
          d.lib_id, d.lib_time, d.lib_comment
          from sample_info_detail d inner join sample_project_master m on m.id=d.project_id""",
            'upmachine': """select d.id, d.sample_id, d.sample_name,
            d.upmachine_type, d.upmachine_mode, d.upmachine_num, d.upmachine_time, d.upmachine_comment
          from sample_info_detail d inner join sample_project_master m on m.id=d.project_id""",
            'downmachine': """select d.id, d.sample_id, d.sample_name,
          d.downmachine_num, d.q20, d.q30, d.downmachine_time, d.downmachine_comment
          from sample_info_detail d inner join sample_project_master m on m.id=d.project_id"""}

get_summary_cmd = '''
      select ss.sample_name,ss.sample_id,
      ss.om_id,qc.results,um.data_count,
      dm.data_count,ss.location,um.location from
      send_sample ss left join quality_check qc on
      ss.sample_id=qc.sample_id left join upmachine um on
      qc.sample_id=um.sample_id left join downmachine dm on
      um.sample_id=dm.sample_id
      where ss.project_id='{project_id}' and
      (qc.project_id='{project_id}' or qc.project_id is NULL) and
      (um.project_id='{project_id}' or um.project_id is NULL) and
      (dm.project_id='{project_id}' or dm.project_id is NULL) and
      ss.status='Y' and
      (qc.status='Y' or qc.status is NULL) and
      (um.status='Y' or um.status is NULL) and
      (dm.status='Y' or dm.status is NULL);'''


def get_proj_name_by_id(project_id):
    cmd = "SELECT project_number FROM sample_project_master where id='%s'" % project_id
    cursor = connection.cursor()
    cursor.execute(cmd)
    result = cursor.fetchone()

    return result[0] if result else u"请选择项目"


def get_all_proj_info():
    cmd = "select id, project_number from sample_project_master"
    cursor = connection.cursor()
    cursor.execute(cmd)
    result = cursor.fetchall()

    return result


def get_sample_by_project(project_id, name):
    """
    :param project_id:
    :return:
    where project_id='%s'
    """
    cmd = "select * from %s where project_id='%s' and status='Y'" % (name, project_id)

    cursor = connection.cursor()
    cursor.execute(cmd)
    result = cursor.fetchall()

    return result


def get_user_id_by_name(username):
    cmd = "select id from user_info where username='%s'" % username
    cursor = connection.cursor()
    cursor.execute(cmd)
    result = cursor.fetchone()

    return result[0] if result else 0


def get_upload_time(project_id, name):
    if name == 'return_sample':
        cmd = "select distinct upload_time,status from %s where status='Y' order by upload_time desc" % (name)
    else:
        cmd = "select distinct upload_time,status from %s where project_id='%s' and status='Y' order by upload_time desc" % (name, project_id)
    cursor = connection.cursor()
    cursor.execute(cmd)
    results = cursor.fetchall()

    return results


def get_return_sample():
    cmd = "select * from return_sample where status='Y'"

    cursor = connection.cursor()
    cursor.execute(cmd)
    results = cursor.fetchall()

    return results


def get_project_summary(project_id):
    cmd = get_summary_cmd.format(project_id=project_id)
    cursor = connection.cursor()
    cursor.execute(cmd)
    results = cursor.fetchall()
    return results


def download_summary_table(project_id):
    title_list = [u'样品名称', u'样品编号', 'OMID', u'质检结果',
                  u'上机数据量', u'下机数据量', u'样品所在地', u'项目所在地']
    book = Workbook()
    sheet1 = book.worksheets[0]
    cursor = connection.cursor()
    cmd = get_summary_cmd.format(project_id=project_id)
    cursor.execute(cmd)

    results = cursor.fetchall()

    for i, title in enumerate(title_list):
        sheet1.cell(row=1, column=i+1).value = title
    for i, result in enumerate(results):
        for j, data in enumerate(result):
            sheet1.cell(row=i+2, column=j+1).value = data

    export_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'export')

    if not os.path.exists(export_path):
        os.mkdir(export_path)

    file_name = '_'.join(['summary', str(project_id)]) + '_' + datetime.datetime.now().strftime("%Y%m%d%H%M%S") + '.xls'
    full_file_name = os.path.join(export_path, file_name)
    book.save(filename=full_file_name)
    return full_file_name
