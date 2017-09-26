# coding: utf8
import os
import sys
import datetime
from math import ceil
from openpyxl import Workbook
from django.db import connection
from ..models import BillingInfo, CostInfo, ReceiptInfo
'''
deal with user cost functions
'''

cmd_dict = {'billing_info': '''select bi.* ,sm.cust_user from billing_info bi
                               inner join sample_project_master sm on bi.project_id=sm.id
                               where sm.cust_user="%s"''',
           'receipt_info': '''select rt.* ,sm.cust_user from receipt_info rt
                              inner join sample_project_master sm on rt.project_id=sm.id
                              where sm.cust_user="%s"''',
           'cost_info': '''select ct.* ,sm.cust_user from cost_info ct
                           inner join sample_project_master sm on ct.project_id=sm.id
                           where sm.cust_user="%s"'''}

billing_map = {u'项目编号': 'project_number',
               u'金额': 'expense',
               u'开票': 'billing_number',
               u'时间': 'time',
               u'备注': 'comment'}
receipt_map = {u'项目编号': 'project_number',
               u'金额': 'expense',
               u'时间': 'time',
               u'备注': 'comment'}
cost_map = {u'项目编号': 'project_number',
            u'金额': 'expense',
            u'样品数量': 'sample_number',
            u'单价': 'unit_cost',
            u'时间': 'time',
            u'备注': 'comment'}
title_map = {'billing_info': billing_map,
             'receipt_info': receipt_map,
             'cost_info': cost_map}
title_list = {'billing_info': [u'项目编号', u'金额', u'开票', u'时间', u'备注'],
              'receipt_info': [u'项目编号', u'金额', u'时间', u'备注'],
              'cost_info': [u'项目编号', u'金额', u'样品数量', u'单价', u'时间', u'备注']}


def get_expense_info(user, limit, sidx, sord, page, table):
    cmd1 = cmd_dict[table] % (user)
    cursor = connection.cursor()
    cursor.execute(cmd1)
    results = cursor.fetchall()
    count = len(results)
    if count > 0:
        total_pages = ceil(float(count)/float(limit))
    else:
        total_pages = 0

    if page > total_pages:
        page = total_pages
    if page < 1:
        start = 0
    else:
        start = int(limit) * page - int(limit)
    cmd2 = cmd1 + " order by {sort_index} {sort_direct} limit {start}, {limit}".format(sort_index=sidx,
                                                                                       sort_direct=sord,
                                                                                       start=int(start),
                                                                                       limit=limit)
    cursor = connection.cursor()
    cursor.execute(cmd2)
    results = cursor.fetchall()

    return page, total_pages, count, results


def get_expense_total(user, table, project_id=''):
    if not project_id:
        cmd = cmd_dict[table] % (user)
    else:
        cmd = cmd_dict[table] % (user)
        cmd = cmd + " and project_id='%s'" % (project_id)
    # print cmd
    cursor = connection.cursor()
    cursor.execute(cmd)

    results = cursor.fetchall()
    total_expense = 0
    if results:
        for result in results:
            total_expense += result[3]
        return total_expense
    else:
        return total_expense


def check_bill(user, project_id):
    bi_exp = get_expense_total(user, 'billing_info', project_id)
    re_exp = get_expense_total(user, 'receipt_info', project_id)
    co_exp = get_expense_total(user, 'cost_info', project_id)

    exp_dict = {'billing_info': bi_exp,
                'receipt_info': re_exp,
                'cost_info': co_exp}
    if bi_exp == re_exp == co_exp == 0:
        exp_dict.update({'status': 'ok'})
    elif bi_exp > co_exp and re_exp >= bi_exp * 0.6 and re_exp > co_exp:
        exp_dict.update({'status': 'ok'})
    else:
        exp_dict.update({'status': 'bad'})
    return exp_dict


def down_expense_table(user, table):
    file_title = {'billing_info': u'开票信息表', 'receipt_info': u'收款信息表', 'cost_info': u'成本信息表'}
    book = Workbook()
    sheet1 = book.worksheets[0]
    cmd = cmd_dict[table] % (user)
    cursor = connection.cursor()
    cursor.execute(cmd)

    results = cursor.fetchall()
    results = [result[2:len(result)-1] for result in results]

    for i, title in enumerate(title_list[table]):
        sheet1.cell(row=1, column=i+1).value = title
    for i, result in enumerate(results):
        for j, data in enumerate(result):
            sheet1.cell(row=i+2, column=j+1).value = data

    export_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'export')

    if not os.path.exists(export_path):
        os.mkdir(export_path)

    file_name = '_'.join([user, file_title[table]]) + datetime.datetime.now().strftime("%Y%m%d%H%M%S") + '.xls'
    full_file_name = os.path.join(export_path, file_name)
    book.save(filename=full_file_name)
    return full_file_name
